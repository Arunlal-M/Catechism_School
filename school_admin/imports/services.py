"""
Excel import service — the business logic for processing uploaded files.

WHY A SERVICE LAYER?
────────────────────
Import logic is complex: read Excel, map columns, validate rows, detect
duplicates, create records, track failures. Putting all this in a view
would make the view 500+ lines and untestable.

A service module contains plain Python functions/classes that:
  - Take data in, return results
  - Don't know about HTTP requests or templates
  - Are easy to test with unit tests
  - Can be called from views, management commands, or APIs

THIS IS THE MOST IMPORTANT ARCHITECTURE PATTERN IN THE PROJECT.
"""

from datetime import datetime

from django.db import transaction

from admissions.models import Division, Provision, SchoolClass, Unit, Occupation
from common.constants import (
    IMPORT_COMPLETED,
    IMPORT_FAILED,
    IMPORT_PROCESSING,
    ROW_DUPLICATE,
    ROW_FAILED,
    ROW_SUCCESS,
    STATUS_PENDING,
)
from students.models import CatechismInfo, ParentInfo, StudentProfile

from .models import AdmissionImportBatch, ImportRow

import openpyxl


# ──────────────────────────────────────────────
# Column mapping: user-friendly names → DB field paths
# ──────────────────────────────────────────────
IMPORTABLE_FIELDS = {
    "student_name": "Student Name",
    "email": "Email",
    "gender": "Gender (M/F/O)",
    "date_of_birth": "Date of Birth",
    "school_class": "Class",
    "division": "Division",
    "residential_address": "Residential Address",
    "attended_catechism": "Attended Catechism (Yes/No)",
    "remarks": "Remarks",
    "father_name": "Father's Name",
    "father_occupation": "Father's Occupation",
    "father_phone": "Father's Phone",
    "mother_name": "Mother's Name",
    "mother_occupation": "Mother's Occupation",
    "mother_phone": "Mother's Phone",
    "catechism_teacher_name": "Catechism Teacher Name",
    "provision": "Provision / Pradeshikam",
    "unit": "Unit",
}


def read_excel_headers(file_path):
    """
    Read the first row of an Excel file to get column headers.

    Returns: list of header strings, e.g., ["Name", "Email", "Class", ...]
    """
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value).strip() if cell.value else "")
    wb.close()
    return headers


def auto_suggest_mapping(excel_headers):
    """
    Try to auto-match Excel headers to database fields.

    HOW: Normalize both sides to lowercase and check for containment.
    e.g., "Student Name" matches "student_name" because both contain "student" and "name".
    """
    suggestions = {}
    db_fields_lower = {k: k for k in IMPORTABLE_FIELDS}

    for header in excel_headers:
        header_lower = header.lower().replace(" ", "_").replace("'", "").replace("'", "")
        # Direct match
        if header_lower in db_fields_lower:
            suggestions[header] = header_lower
            continue
        # Fuzzy match: check if the header contains a db field keyword
        for db_field in db_fields_lower:
            # Remove common prefixes and check
            if db_field.replace("_", "") in header_lower.replace("_", ""):
                suggestions[header] = db_field
                break

    return suggestions


def process_import(batch_id, user):
    """
    Main import function — processes all rows in a batch.

    Steps:
    1. Read the Excel file
    2. Apply column mapping
    3. Validate each row
    4. Check for duplicates
    5. Create student records
    6. Update batch counters
    """
    batch = AdmissionImportBatch.objects.get(pk=batch_id)
    batch.status = IMPORT_PROCESSING
    batch.save()

    try:
        wb = openpyxl.load_workbook(batch.file.path)
        ws = wb.active

        # Get headers from row 1
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

        # Build column index map: {db_field: column_index}
        mapping = batch.column_mapping  # {"Excel Header": "db_field"}
        field_to_col = {}
        for excel_header, db_field in mapping.items():
            if excel_header in headers:
                field_to_col[db_field] = headers.index(excel_header)

        row_count = 0
        success = 0
        failed = 0
        duplicates = 0

        # Process data rows (skip header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue

            row_count += 1
            raw_data = {headers[i]: str(v) if v else "" for i, v in enumerate(row) if i < len(headers)}

            # Extract values using the mapping
            data = {}
            for db_field, col_idx in field_to_col.items():
                if col_idx < len(row):
                    data[db_field] = row[col_idx]

            # Validate and create
            result = _process_single_row(data, user)

            import_row = ImportRow(
                batch=batch,
                row_number=row_idx,
                raw_data=raw_data,
                status=result["status"],
                error_message=result.get("error", ""),
                student=result.get("student"),
            )
            import_row.save()

            if result["status"] == ROW_SUCCESS:
                success += 1
            elif result["status"] == ROW_DUPLICATE:
                duplicates += 1
            else:
                failed += 1

        wb.close()

        # Update batch summary
        batch.total_rows = row_count
        batch.success_count = success
        batch.failed_count = failed
        batch.duplicate_count = duplicates
        batch.status = IMPORT_COMPLETED
        batch.save()

    except Exception as e:
        batch.status = IMPORT_FAILED
        batch.save()
        raise e


def _process_single_row(data, user):
    """
    Validate one row and create student record if valid.

    Returns: {"status": "success|failed|duplicate", "error": "...", "student": obj}
    """
    errors = []

    # ── Required field validation ──
    student_name = str(data.get("student_name", "")).strip()
    if not student_name:
        errors.append("Student name is required.")

    # ── Parse date of birth ──
    dob = data.get("date_of_birth")
    if dob:
        if isinstance(dob, datetime):
            dob = dob.date()
        elif isinstance(dob, str):
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
                try:
                    dob = datetime.strptime(dob.strip(), fmt).date()
                    break
                except ValueError:
                    continue
            else:
                errors.append(f"Invalid date format: {dob}")
                dob = None
    else:
        errors.append("Date of birth is required.")

    # ── Gender validation ──
    gender_raw = str(data.get("gender", "")).strip().upper()
    gender_map = {"M": "M", "F": "F", "O": "O", "MALE": "M", "FEMALE": "F", "OTHER": "O"}
    gender = gender_map.get(gender_raw)
    if not gender:
        errors.append(f"Invalid gender: {data.get('gender')}. Use M, F, or O.")

    # ── Resolve class (lookup) ──
    class_name = str(data.get("school_class", "")).strip()
    school_class = None
    if class_name:
        school_class = SchoolClass.objects.filter(name__iexact=class_name).first()
        if not school_class:
            errors.append(f"Class '{class_name}' not found. Add it in Master Data first.")
    else:
        errors.append("Class is required.")

    # ── Resolve division (lookup) ──
    division_name = str(data.get("division", "")).strip()
    division = None
    if division_name:
        division = Division.objects.filter(name__iexact=division_name).first()
        if not division:
            errors.append(f"Division '{division_name}' not found. Add it in Master Data first.")
    else:
        errors.append("Division is required.")

    # ── Resolve occupations (lookup or create) ──
    father_occ_name = str(data.get("father_occupation", "")).strip()
    father_occupation = None
    if father_occ_name:
        father_occupation, _ = Occupation.objects.get_or_create(name__iexact=father_occ_name, defaults={"name": father_occ_name})

    mother_occ_name = str(data.get("mother_occupation", "")).strip()
    mother_occupation = None
    if mother_occ_name:
        mother_occupation, _ = Occupation.objects.get_or_create(name__iexact=mother_occ_name, defaults={"name": mother_occ_name})

    # ── If validation failed, stop here ──
    if errors:
        return {"status": ROW_FAILED, "error": " | ".join(errors)}

    # ── Duplicate detection ──
    # Rule: match on name + father's phone + DOB
    father_phone = str(data.get("father_phone", "")).strip()
    existing = StudentProfile.objects.filter(
        student_name__iexact=student_name,
        date_of_birth=dob,
    )
    if father_phone:
        existing = existing.filter(parent_info__father_phone=father_phone)

    if existing.exists():
        return {"status": ROW_DUPLICATE, "error": f"Duplicate: student '{student_name}' already exists."}

    # ── Create the records ──
    try:
        with transaction.atomic():
            # Student
            attended = str(data.get("attended_catechism", "")).strip().lower()
            student = StudentProfile.objects.create(
                student_name=student_name,
                email=str(data.get("email", "")).strip(),
                gender=gender,
                date_of_birth=dob,
                school_class=school_class,
                division=division,
                residential_address=str(data.get("residential_address", "")).strip(),
                attended_catechism=attended in ("yes", "y", "true", "1"),
                status=STATUS_PENDING,
                remarks=str(data.get("remarks", "")).strip(),
                created_by=user,
                updated_by=user,
            )

            # Parent info
            ParentInfo.objects.create(
                student=student,
                father_name=str(data.get("father_name", "")).strip(),
                father_occupation=father_occupation,
                father_phone=father_phone,
                mother_name=str(data.get("mother_name", "")).strip(),
                mother_occupation=mother_occupation,
                mother_phone=str(data.get("mother_phone", "")).strip(),
            )

            # Catechism info
            provision = None
            provision_name = str(data.get("provision", "")).strip()
            if provision_name:
                provision = Provision.objects.filter(name__iexact=provision_name).first()

            unit = None
            unit_name = str(data.get("unit", "")).strip()
            if unit_name and provision:
                unit = Unit.objects.filter(name__iexact=unit_name, provision=provision).first()

            CatechismInfo.objects.create(
                student=student,
                catechism_teacher_name=str(data.get("catechism_teacher_name", "")).strip(),
                provision=provision,
                unit=unit,
            )

        return {"status": ROW_SUCCESS, "student": student}

    except Exception as e:
        return {"status": ROW_FAILED, "error": str(e)}
