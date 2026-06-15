"""
Export service — CSV and Excel export of filtered student data.

WHY A SEPARATE EXPORTER?
────────────────────────
Same principle as the import service: keep business logic out of views.
Views handle HTTP; exporters handle data formatting.
"""

import csv
import io
from datetime import datetime

import openpyxl
from django.http import HttpResponse


def export_students_csv(queryset):
    """
    Export a filtered queryset of students as a CSV download.

    HOW StreamingHttpResponse WORKS:
    For large datasets, we could use StreamingHttpResponse to avoid
    loading everything into memory. For school-sized data (hundreds,
    not millions), a regular HttpResponse with StringIO is fine.
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Header row
    writer.writerow([
        "Student Number", "Student Name", "Email", "Gender", "Date of Birth",
        "Class", "Division", "Address", "Attended Catechism",
        "Father Name", "Father Occupation", "Father Phone",
        "Mother Name", "Mother Occupation", "Mother Phone",
        "Catechism Teacher", "Provision", "Unit",
        "Status", "Remarks",
    ])

    # Data rows
    for student in queryset.select_related(
        "school_class", "division", "parent_info",
        "catechism_info", "catechism_info__provision", "catechism_info__unit",
    ):
        parent = getattr(student, "parent_info", None)
        catechism = getattr(student, "catechism_info", None)

        writer.writerow([
            student.student_number,
            student.student_name,
            student.email,
            student.get_gender_display(),
            student.date_of_birth.strftime("%d-%m-%Y") if student.date_of_birth else "",
            str(student.school_class) if student.school_class else "",
            str(student.division) if student.division else "",
            student.residential_address,
            "Yes" if student.attended_catechism else "No",
            parent.father_name if parent else "",
            str(parent.father_occupation) if parent and parent.father_occupation else "",
            parent.father_phone if parent else "",
            parent.mother_name if parent else "",
            str(parent.mother_occupation) if parent and parent.mother_occupation else "",
            parent.mother_phone if parent else "",
            catechism.catechism_teacher_name if catechism else "",
            str(catechism.provision) if catechism and catechism.provision else "",
            str(catechism.unit) if catechism and catechism.unit else "",
            student.get_status_display(),
            student.remarks,
        ])

    response = HttpResponse(output.getvalue(), content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="students_export_{datetime.now():%Y%m%d_%H%M}.csv"'
    return response


def export_students_excel(queryset):
    """Export a filtered queryset as an Excel (.xlsx) download."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    # Header row (bold)
    headers = [
        "Student Number", "Student Name", "Email", "Gender", "Date of Birth",
        "Class", "Division", "Address", "Attended Catechism",
        "Father Name", "Father Occupation", "Father Phone",
        "Mother Name", "Mother Occupation", "Mother Phone",
        "Catechism Teacher", "Provision", "Unit",
        "Status", "Remarks",
    ]
    ws.append(headers)

    # Bold headers
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data rows
    for student in queryset.select_related(
        "school_class", "division", "parent_info",
        "catechism_info", "catechism_info__provision", "catechism_info__unit",
    ):
        parent = getattr(student, "parent_info", None)
        catechism = getattr(student, "catechism_info", None)

        ws.append([
            student.student_number,
            student.student_name,
            student.email,
            student.get_gender_display(),
            student.date_of_birth.strftime("%d-%m-%Y") if student.date_of_birth else "",
            str(student.school_class) if student.school_class else "",
            str(student.division) if student.division else "",
            student.residential_address,
            "Yes" if student.attended_catechism else "No",
            parent.father_name if parent else "",
            str(parent.father_occupation) if parent and parent.father_occupation else "",
            parent.father_phone if parent else "",
            parent.mother_name if parent else "",
            str(parent.mother_occupation) if parent and parent.mother_occupation else "",
            parent.mother_phone if parent else "",
            catechism.catechism_teacher_name if catechism else "",
            str(catechism.provision) if catechism and catechism.provision else "",
            str(catechism.unit) if catechism and catechism.unit else "",
            student.get_status_display(),
            student.remarks,
        ])

    # Auto-width columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    # Write to response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="students_export_{datetime.now():%Y%m%d_%H%M}.xlsx"'
    return response


def export_failed_rows_excel(batch):
    """Export failed import rows as an Excel file for re-upload."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Failed Rows"

    failed_rows = batch.rows.filter(status="failed").order_by("row_number")
    if not failed_rows.exists():
        ws.append(["No failed rows"])
    else:
        # Use the first row's keys as headers + add error column
        first_row = failed_rows.first()
        headers = list(first_row.raw_data.keys()) + ["Error"]
        ws.append(headers)

        from openpyxl.styles import Font
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for row in failed_rows:
            values = [row.raw_data.get(h, "") for h in headers[:-1]]
            values.append(row.error_message)
            ws.append(values)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="failed_rows_batch_{batch.pk}.xlsx"'
    return response
